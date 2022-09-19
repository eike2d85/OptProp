from airfoil_sim_parameters import airfoil_sim_parameters
import numpy as np
import openpyxl as xl

Re_min,Re_max,Re_step,alpha_i,alpha_f,alpha_step,n_iter,airfoil_list = airfoil_sim_parameters()

H_bisco_Re = np.arange(Re_min, Re_max, Re_step)
H_bisco_alpha = np.arange(alpha_i,alpha_f,alpha_step)


    
def airfoil_data_get(nperfil, alpha, Re):
    alpha = alpha*180/np.pi
    if Re<Re_min:
        Re=Re_min
        print('Reynolds menor que banco de dados')
    if Re>Re_max:
        Re=Re_max
        print('Reynolds maior que banco de dados')
    j=0
    for Rey in H_bisco_Re:
        if Re >= Rey:
            Rey_maior = Rey + Re_step
            Rey_menor = Rey
            for aoa in H_bisco_alpha:
                if alpha >= aoa:
                    aoa_maior = aoa+alpha_step
                    aoa_menor = aoa 
                    j_save = j 
                j=j+1
            j=0
            
        

    workbook = xl.load_workbook(nperfil+'.xlsx', "rb")
    worksheet = workbook[str(Rey_menor)]
    Cl_aoa_menor = worksheet.cell(j_save+1,2).value
    Cd_aoa_menor = worksheet.cell(j_save+1,3).value
    Cdp_aoa_menor = worksheet.cell(j_save+1,4).value
    Cl_aoa_maior = worksheet.cell(j_save+2,2).value
    Cd_aoa_maior = worksheet.cell(j_save+2,3).value
    Cdp_aoa_maior = worksheet.cell(j_save+2,4).value

    Cl_Re_menor = np.interp(alpha,[aoa_menor,aoa_maior],[Cl_aoa_menor,Cl_aoa_maior]) 
    Cd_Re_menor = np.interp(alpha,[aoa_menor,aoa_maior],[Cd_aoa_menor,Cd_aoa_maior]) 
    Cdp_Re_menor = np.interp(alpha,[aoa_menor,aoa_maior],[Cdp_aoa_menor,Cdp_aoa_maior]) 

    worksheet = workbook[str(Rey_maior)]
    Cl_aoa_menor = worksheet.cell(j_save+1,2).value
    Cd_aoa_menor = worksheet.cell(j_save+1,3).value
    Cdp_aoa_menor = worksheet.cell(j_save+1,4).value
    Cl_aoa_maior = worksheet.cell(j_save+2,2).value
    Cd_aoa_maior = worksheet.cell(j_save+2,3).value
    Cdp_aoa_maior = worksheet.cell(j_save+2,4).value

    Cl_Re_maior = np.interp(alpha,[aoa_menor,aoa_maior],[Cl_aoa_menor,Cl_aoa_maior]) 
    Cd_Re_maior = np.interp(alpha,[aoa_menor,aoa_maior],[Cd_aoa_menor,Cd_aoa_maior]) 
    Cdp_Re_maior = np.interp(alpha,[aoa_menor,aoa_maior],[Cdp_aoa_menor,Cdp_aoa_maior]) 


    Cl = np.interp(Re,[Rey_menor,Rey_maior],[Cl_Re_menor,Cl_Re_maior]) 
    Cd = np.interp(Re,[Rey_menor,Rey_maior],[Cd_Re_menor,Cd_Re_maior]) 
    Cdp = np.interp(Re,[Rey_menor,Rey_maior],[Cdp_Re_menor,Cdp_Re_maior]) 
    return Cl,Cd,Cdp

#x, y, z = airfoil_data_get('NACA2412', 6.37, 517000)
#print(x,y,z)